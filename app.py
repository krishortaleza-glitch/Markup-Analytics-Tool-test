import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

st.set_page_config(page_title="Wholesale Markup Analytics", layout="wide")
st.title("💰 Wholesale Markup Analytics Tool")

@st.cache_data
def load_file(file):
    if file.name.endswith(".csv"):
        return pd.read_csv(file)
    return pd.read_excel(file)

st.header("Upload Files")

inv_file = st.file_uploader("Invoices")
prod_file = st.file_uploader("Products File")
front_file = st.file_uploader("Frontline")
tax_file = st.file_uploader("Taxes")
store_file = st.file_uploader("Storelist")

st.markdown("### 📊 Markup % Formula")
st.info("Markup % = (Invoice Cost - (Frontline + Tax)) / (Frontline + Tax)")

if inv_file and prod_file and front_file and tax_file and store_file:

    inv = load_file(inv_file)
    prod = load_file(prod_file)
    front = load_file(front_file)
    tax = load_file(tax_file)
    store = load_file(store_file)

    st.success("Files loaded")

    # ==============================
    # COLUMN MAPPING
    # ==============================
    inv_store = "store"
    inv_product = "productId"
    inv_cost = "price"

    prod_id = "ProductId"
    prod_family = "Family"
    prod_type = "Type"
    prod_case = "Products/Case"

    front_family = "Family"
    front_cost = "CasePrice"
    front_start = "Start"
    front_end = "End"

    store_store = "uniqueId"
    store_state = "stateAbbrev"

    tax_state = st.selectbox("Tax State", tax.columns)
    tax_value = st.selectbox("Tax Value", tax.columns)

    if st.button("🚀 Run Analysis"):

        progress = st.progress(0)

        # ==============================
        # CLEAN KEYS
        # ==============================
        def clean_id(x):
            return str(x).strip().lstrip("0")

        inv["ProductID"] = inv[inv_product].apply(clean_id)
        prod["ProductID"] = prod[prod_id].apply(clean_id)

        inv["Store"] = inv[inv_store].astype(str).str.strip()
        store["Store"] = store[store_store].astype(str).str.strip()

        store["State"] = store[store_state].astype(str).str.strip()
        tax["State"] = tax[tax_state].astype(str).str.strip()

        prod["Family"] = prod[prod_family].astype(str).str.strip().str.upper()
        front["Family"] = front[front_family].astype(str).str.strip().str.upper()

        progress.progress(10)

        # ==============================
        # 🔥 FORCE PRODUCT UNIQUENESS
        # ==============================
        prod = (
            prod
            .sort_values(prod_case, ascending=False)
            .drop_duplicates(subset=["ProductID"], keep="first")
        )

        # ==============================
        # MERGE PRODUCT
        # ==============================
        merged = inv.merge(
            prod[["ProductID", "Family", "Type", prod_case]],
            on="ProductID",
            how="left"
        )

        progress.progress(25)

        # ==============================
        # ACTIVE FRONTLINE
        # ==============================
        today = pd.Timestamp.today().normalize()

        front[front_start] = pd.to_datetime(front[front_start], errors="coerce")
        front[front_end] = pd.to_datetime(front[front_end], errors="coerce")
        front[front_end] = front[front_end].fillna(pd.Timestamp.max)

        active_front = front[
            (front[front_start] <= today) &
            (front[front_end] >= today)
        ].copy()

        active_front = (
            active_front
            .sort_values(front_start, ascending=False)
            .drop_duplicates(subset=["Family"])
        )

        progress.progress(45)

        # ==============================
        # MERGE FRONTLINE
        # ==============================
        merged = merged.merge(
            active_front[["Family", front_cost]],
            on="Family",
            how="left"
        )

        progress.progress(60)

        # ==============================
        # STORE → STATE
        # ==============================
        merged = merged.merge(
            store[["Store", "State"]],
            on="Store",
            how="left"
        )

        progress.progress(75)

        # ==============================
        # TAX MERGE
        # ==============================
        merged = merged.merge(
            tax[["State", tax_value]],
            on="State",
            how="left"
        )

        progress.progress(85)

        # ==============================
        # TAX RULE ENGINE
        # ==============================
        merged["Base Tax"] = pd.to_numeric(merged[tax_value], errors="coerce")
        merged["Frontline"] = pd.to_numeric(merged[front_cost], errors="coerce")
        merged[prod_case] = pd.to_numeric(merged[prod_case], errors="coerce")

        merged["Tax"] = merged["Base Tax"]
        merged["Tax Rule Applied"] = "Default"

        merged["Type"] = merged["Type"].astype(str).str.strip()
        merged["State"] = merged["State"].astype(str).str.strip().str.upper()

        # TX
        mask_tx = (
            merged["Type"].isin(["Modern Oral", "Smokeless", "Smokeless Big", "Snus"]) &
            (merged["State"] == "TX")
        )
        merged.loc[mask_tx, "Tax"] = merged[prod_case] * merged["Base Tax"]
        merged.loc[mask_tx, "Tax Rule Applied"] = "TX: Case * Tax"

        # CO Smokeless Big
        mask_co_big = (
            (merged["Type"] == "Smokeless Big") &
            (merged["State"] == "CO")
        )
        merged.loc[mask_co_big, "Tax"] = merged[prod_case] * merged["Base Tax"]
        merged.loc[mask_co_big, "Tax Rule Applied"] = "CO Big: Case * Tax"

        # KS, NM
        mask_ks_nm = (
            merged["Type"].isin(["Modern Oral", "Smokeless", "Smokeless Big", "Snus"]) &
            merged["State"].isin(["KS", "NM"])
        )
        merged.loc[mask_ks_nm, "Tax"] = (merged["Base Tax"] / 100) * merged["Frontline"]
        merged.loc[mask_ks_nm, "Tax Rule Applied"] = "KS/NM: % * Frontline"

        # CO %
        mask_co_pct = (
            merged["Type"].isin(["Modern Oral", "Smokeless", "Snus"]) &
            (merged["State"] == "CO")
        )
        merged.loc[mask_co_pct, "Tax"] = (merged["Base Tax"] / 100) * merged["Frontline"]
        merged.loc[mask_co_pct, "Tax Rule Applied"] = "CO: % * Frontline"

        merged["Tax"] = merged["Tax"].fillna(0)

        # ==============================
        # 🔥 HARD DEDUP (FINAL FIX)
        # ==============================
        merged = merged.sort_values("Tax", ascending=False)

        merged = merged.drop_duplicates(
            subset=["State", "Family", "Type", "Invoice Cost"],
            keep="first"
        )

        # ==============================
        # CALCULATIONS
        # ==============================
        merged["Invoice Cost"] = pd.to_numeric(merged[inv_cost], errors="coerce")

        merged["Total Cost"] = merged["Frontline"] + merged["Tax"]
        merged["Markup"] = merged["Invoice Cost"] - merged["Total Cost"]
        merged["Markup %"] = merged["Markup"] / merged["Total Cost"]

        merged["Markup %"] = merged["Markup %"].replace([float("inf"), -float("inf")], 0)

        merged["Total Cost"] = merged["Total Cost"].round(2)
        merged["Markup"] = merged["Markup"].round(2)
        merged["Markup %"] = merged["Markup %"].round(3)

        progress.progress(90)

        # ==============================
        # FREQUENCY (FIXED)
        # ==============================
        freq = (
            merged
            .groupby(["State", "Family", "Type", "Invoice Cost"])
            .size()
            .reset_index(name="Frequency")
        )

        freq["Top"] = (
            freq.groupby(["State", "Family", "Type"])["Frequency"]
            .transform("max") == freq["Frequency"]
        )

        merged = merged.merge(freq, on=["State", "Family", "Type", "Invoice Cost"], how="left")

        # ==============================
        # FINAL OUTPUT
        # ==============================
        final = merged[[
            "State","Family","Type","Invoice Cost","Frontline","Tax",
            "Total Cost","Markup","Markup %","Frequency","Top","Tax Rule Applied"
        ]].drop_duplicates()

        # ==============================
        # EXPORT
        # ==============================
        output = BytesIO()

        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            final.to_excel(writer, sheet_name="Analysis", index=False)
            merged.to_excel(writer, sheet_name="Full Output", index=False)

        output.seek(0)

        wb = load_workbook(output)
        ws = wb["Analysis"]

        green = PatternFill(start_color="C6EFCE", fill_type="solid")
        top_col = list(final.columns).index("Top") + 1

        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row=row_idx, column=top_col).value:
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = green

        final_output = BytesIO()
        wb.save(final_output)
        final_output.seek(0)

        progress.progress(100)

        st.download_button(
            "📥 Download Analysis",
            data=final_output,
            file_name=f"markup_analysis_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        )
